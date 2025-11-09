from flask import Flask, render_template_string, request, send_file
import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference

app = Flask(__name__)
UPLOAD_FOLDER = "uploads"
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

# -------------------- HTML PAGE --------------------
HTML_PAGE = """
<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Excel Report Generator</title>
<style>
body {
    font-family: 'Segoe UI', sans-serif;
    background: linear-gradient(135deg, #1a2a5a 0%, #1e3c72 25%, #00c6ff 50%, #1e3c72 75%, #1a2a5a 100%);
    height: 100vh;
    margin: 0;
    display: flex;
    align-items: center;
    justify-content: center;
    overflow: hidden;
}
body::before {
    content: "";
    position: absolute;
    inset: 0;
    background: rgba(0,0,0,0.45);
    backdrop-filter: blur(6px);
    z-index: 0;
}
.container {
    background-color: #0d1a3a;
    border-radius: 20px;
    box-shadow: 0 8px 30px rgba(0,0,0,0.4);
    padding: 40px;
    text-align: center;
    width: 400px;
    position: relative;
    z-index: 1;
    backdrop-filter: blur(15px);
    border: 1px solid rgba(255,255,255,0.3);
}
h2 {
    background: linear-gradient(90deg, #00c6ff, #007bff);
    -webkit-background-clip: text;
    -webkit-text-fill-color: transparent;
    font-size: 26px;
    margin-bottom: 20px;
    font-weight: 700;
}
.upload-box {
    border: 2px dashed rgba(255, 255, 255, 0.7);
    padding: 20px;
    border-radius: 12px;
    margin-bottom: 25px;
    color: white;
    cursor: pointer;
    transition: 0.3s;
}
.upload-box:hover { background: rgba(255,255,255,0.1); transform: scale(1.02);}
input[type=file] { display: none; }
.btn {
    background: linear-gradient(135deg, #007bff, #00c6ff);
    color: white;
    border: none;
    padding: 12px 20px;
    margin: 10px 0;
    border-radius: 8px;
    cursor: pointer;
    width: 80%;
    font-size: 15px;
    font-weight: 600;
    transition: 0.3s;
}
.btn:hover { transform: scale(1.03); }
.status { margin-top: 10px; color: #fff; font-weight: 600; }
.footer { margin-top: 15px; color: rgba(255,255,255,0.8); font-size: 13px; font-weight: 500; }
</style>
</head>
<body>
<div class="container">
<h2>Excel Report Generator</h2>
<div class="upload-box" onclick="document.getElementById('fileUpload').click();">
    📂 Choose Excel File
</div>
<input id="fileUpload" type="file" accept=".xlsx">
<button class="btn" onclick="uploadReport('participation')">Generate Participation Report</button>
<button class="btn" onclick="uploadReport('performance')">Generate Performance Report</button>
<div class="status" id="status"></div>
<div class="footer">Designed with 💙 by Sabapathi</div>
</div>

<script>
let selectedFile = null;
document.getElementById('fileUpload').addEventListener('change', (e) => {
    selectedFile = e.target.files[0];
    document.getElementById('status').innerText = selectedFile ? "📄 File selected: " + selectedFile.name : "";
});

function uploadReport(action) {
    if(!selectedFile) { alert("Please choose a file first!"); return; }
    const status = document.getElementById('status');
    status.innerText = "⏳ Uploading...";

    const formData = new FormData();
    formData.append("file", selectedFile);
    formData.append("action", action);

    fetch("/", { method: "POST", body: formData })
    .then(res => {
        if(!res.ok) throw new Error("Server error");
        return res.blob();
    })
    .then(blob => {
        const url = window.URL.createObjectURL(blob);
        const a = document.createElement("a");
        a.href = url;
        a.download = "Report.xlsx";
        document.body.appendChild(a);
        a.click();
        a.remove();
        status.innerText = "✅ Uploaded & Report Downloaded!";
    })
    .catch(err => { console.error(err); status.innerText = "❌ Error processing file!"; });
}
</script>
</body>
</html>
"""

# -------------------- PARTICIPATION REPORT --------------------
def generate_participation_report(input_file, output_file):
    df = pd.read_excel(input_file)

    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name="Data")

    wb = load_workbook(output_file)
    ws_data = wb["Data"]

    dark_blue_fill = PatternFill(start_color="002060", end_color="002060", fill_type="solid")
    white_font = Font(color="FFFFFF", bold=True)
    border_style = Border(left=Side(style="thin"), right=Side(style="thin"),
                          top=Side(style="thin"), bottom=Side(style="thin"))

    for cell in ws_data[1]:
        cell.fill = dark_blue_fill
        cell.font = white_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in ws_data.iter_rows():
        for cell in row:
            cell.border = border_style
            cell.alignment = Alignment(horizontal="center", vertical="center")

    for col in ws_data.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws_data.column_dimensions[column].width = max_length + 1

    wb.save(output_file)

    pivot_participation = pd.pivot_table(
        df,
        index="Department",
        columns="Test Status",
        values="Name",
        aggfunc="count",
        fill_value=0,
        margins=True,
        margins_name="Grand Total"
    ).reset_index()

    pivot_participation.columns = [str(c) for c in pivot_participation.columns]
    pivot_participation = pivot_participation.loc[:, ~pivot_participation.columns.str.contains('^Unnamed')]

    with pd.ExcelWriter(output_file, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        pivot_participation.to_excel(writer, index=False, sheet_name="Participation Summary")

    wb = load_workbook(output_file)
    ws_summary = wb["Participation Summary"]

    max_row, max_col = ws_summary.max_row, ws_summary.max_column
    header_fill = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    light_blue_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for col in range(1, max_col + 1):
        cell = ws_summary.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in range(2, max_row + 1):
        fill_color = light_blue_fill if row % 2 == 0 else white_fill
        for col in range(1, max_col + 1):
            cell = ws_summary.cell(row=row, column=col)
            cell.fill = fill_color
            cell.border = border_style
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    chart = BarChart()
    chart.title = "Department-wise Test Participation"
    chart.x_axis.title = "Department"
    chart.y_axis.title = "Number of Students"

    data = Reference(ws_summary, min_col=2, max_col=max_col - 1, min_row=1, max_row=max_row - 1)
    categories = Reference(ws_summary, min_col=1, max_col=1, min_row=2, max_row=max_row - 1)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    chart.height = 10
    chart.width = 20

    chart_col = get_column_letter(max_col + 2)
    ws_summary.add_chart(chart, f"{chart_col}2")

    for col in range(1, max_col + 1):
        max_length = 0
        column_letter = get_column_letter(col)
        for row in range(1, max_row + 1):
            val = ws_summary.cell(row=row, column=col).value
            if val:
                max_length = max(max_length, len(str(val)))
        ws_summary.column_dimensions[column_letter].width = max_length + 8

    wb.save(output_file)


# -------------------- PERFORMANCE REPORT --------------------
def generate_performance_report(input_file, output_file):
    df = pd.read_excel(input_file)

    # Participation first
    generate_participation_report(input_file, output_file)

    # ---------------- Find "Test Status" column ----------------
    status_index = None
    for idx, col in enumerate(df.columns):
        if str(col).strip().lower() == "test status":
            status_index = idx
            break
    if status_index is None:
        raise Exception("❌ 'Test Status' column not found in the uploaded file.")

    # ---------------- Find first "total percentage" column after Test Status ----------------
    total_percentage_col = None
    for col in df.columns[status_index + 1:]:
        if "total percentage" in str(col).strip().lower():
            total_percentage_col = col
            break
    if total_percentage_col is None:
        raise Exception("❌ No 'total percentage' column found after 'Test Status'.")

    # ---------------- Categorize function ----------------
    def categorize(percentage):
        if pd.isna(percentage):
            return "Not Attended"
        percentage_str = str(percentage).strip()
        if percentage_str in ["-", "", "NA", "n/a"]:
            return "Not Attended"
        try:
            p = float(percentage_str.replace("%", ""))
        except ValueError:
            return "Not Attended"

        if p > 75:
            return "Good"
        elif p > 50:
            return "Satisfactory"
        elif p > 25:
            return "Need Attention"
        else:
            return "Intervention"

    # ---------------- Insert Category Column ----------------
    df.insert(status_index + 2, "Category", df[total_percentage_col].apply(categorize))

    # ---------------- Pivot Table ----------------
    pivot_perf = pd.pivot_table(
        df,
        index="Department",
        columns="Category",
        values="Name",
        aggfunc="count",
        fill_value=0,
        margins=True,
        margins_name="Grand Total"
    ).reset_index()

    # ---------------- Write Pivot to Excel ----------------
    with pd.ExcelWriter(output_file, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        pivot_perf.to_excel(writer, index=False, sheet_name="Performance Summary")

    # ---------------- Styling and Chart ----------------
    wb = load_workbook(output_file)
    ws_perf = wb["Performance Summary"]

    max_row, max_col = ws_perf.max_row, ws_perf.max_column
    header_fill = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    light_blue_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    border_style = Border(left=Side(style="thin"), right=Side(style="thin"),
                          top=Side(style="thin"), bottom=Side(style="thin"))

    for col in range(1, max_col + 1):
        cell = ws_perf.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in range(2, max_row + 1):
        fill_color = light_blue_fill if row % 2 == 0 else white_fill
        for col in range(1, max_col + 1):
            cell = ws_perf.cell(row=row, column=col)
            cell.fill = fill_color
            cell.border = border_style
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    chart2 = BarChart()
    chart2.title = "Department-wise Performance Summary"
    chart2.x_axis.title = "Department"
    chart2.y_axis.title = "Number of Students"

    data2 = Reference(ws_perf, min_col=2, max_col=max_col - 1, min_row=1, max_row=max_row - 1)
    categories2 = Reference(ws_perf, min_col=1, max_col=1, min_row=2, max_row=max_row - 1)
    chart2.add_data(data2, titles_from_data=True)
    chart2.set_categories(categories2)
    chart2.height = 10
    chart2.width = 20

    chart_col = get_column_letter(max_col + 2)
    ws_perf.add_chart(chart2, f"{chart_col}2")

    for col in range(1, max_col + 1):
        max_length = 0
        column_letter = get_column_letter(col)
        for row in range(1, max_row + 1):
            val = ws_perf.cell(row=row, column=col).value
            if val:
                max_length = max(max_length, len(str(val)))
        ws_perf.column_dimensions[column_letter].width = max_length + 8

    wb.save(output_file)


# -------------------- FLASK ROUTE --------------------
@app.route("/", methods=["GET", "POST"])
def upload_file():
    if request.method == "POST":
        file = request.files.get("file")
        action = request.form.get("action")

        if not file:
            return "⚠️ Please upload an Excel file.", 400

        filepath = os.path.join(UPLOAD_FOLDER, file.filename)
        file.save(filepath)
        output_file = os.path.join(UPLOAD_FOLDER, "Report.xlsx")

        if action == "participation":
            generate_participation_report(filepath, output_file)
        elif action == "performance":
            generate_performance_report(filepath, output_file)
        else:
            return "⚠️ Invalid action.", 400

        return send_file(output_file, as_attachment=True)

    return render_template_string(HTML_PAGE)


if __name__ == "__main__":
    app.run(debug=True, port=5002)
